#!/usr/bin/env python3
"""
rgs_lookup.py — look up and validate RGS (Referentie GrootboekSchema) codes.

Why this exists
---------------
Picking the right RGS reference code (and confirming it actually exists in the version
your software supports) is the single most error-prone step in RGS bookkeeping. This
tool turns the official RGS master into a queryable index so an agent can:
  - validate that a referentiecode exists (optionally within a version filter),
  - look up an account's description, niveau, D/C and omslagcode,
  - fuzzy-search by Dutch description ("debiteuren", "afschrijving", ...),
  - filter to an entity type (BV / EZ / ZZP) and to niveau 4 (the bookable level).

Source of truth
---------------
The authoritative dataset is the official Excel master from
referentiegrootboekschema.nl (Kennisbank > "Download RGS"; current RGS 3.8). Download it
and point this script at it (Excel needs `openpyxl`; a CSV/TSV export needs nothing):

    python rgs_lookup.py --file RGS_3.8.xlsx --lookup BLimKasKas
    python rgs_lookup.py --file RGS_3.8.csv  --search "algemene kosten" --entity BV --nivo 4
    python rgs_lookup.py --file RGS_3.8.xlsx --validate WBedAlkOal

Column headers vary between the official Excel and GBNED/boekhoudplaza exports, so the
loader matches headers case-insensitively against several known aliases. If your export
uses different names, pass them via --col-code / --col-desc etc.

Offline fallback
----------------
With no --file, the script uses a SMALL built-in seed of common, well-attested codes for
a Dutch MKB BV. The seed is a convenience for quick checks ONLY — it is not the complete
schema and must not be treated as authoritative. Always validate against the official
master before booking.

MoneyBird note
--------------
MoneyBird's API is pinned to RGS 3.5 and requires a valid `rgs_code` on ledger-account
creation. Because the RGS core is stable since 3.0, common codes are present in 3.5 — but
if you build a validator for MoneyBird, filter your dataset to codes valid in 3.5.
"""
from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass, field

# --- Header aliases (lowercased) ------------------------------------------------------
# The loader maps each logical field to the first matching column it finds.
HEADER_ALIASES: dict[str, list[str]] = {
    "code": ["referentiecode", "rgs-code", "rgs code", "rgscode", "code", "refcode"],
    "desc": ["omschrijving", "grootboekomschrijving", "standaardomschrijving",
             "omschrijving (lang)", "naam", "description"],
    "nummer": ["referentienummer", "rgs-nr", "rgs nr", "referentiegrootboeknummer",
               "nummer", "reknr", "rgsnr"],
    "nivo": ["niveau", "nivo", "nv", "level"],
    "dc": ["d/c", "dc", "debet/credit", "debet credit", "indicatie d/c"],
    "omslag": ["omslagcode", "omslag", "omslagrekening"],
    "zzp": ["zzp"],
    "ez": ["ez", "ez_vof", "eenmanszaak"],
    "bv": ["bv"],
    "sv": ["sv", "svc", "stichting"],
    "branche": ["branche", "branchecode"],
    "status": ["status", "actief", "vervallen"],
    "sortering": ["sortering", "sorteercode", "sortering bw2"],
}


@dataclass
class RgsAccount:
    code: str
    desc: str = ""
    nummer: str = ""
    nivo: str = ""
    dc: str = ""
    omslag: str = ""
    entities: dict[str, str] = field(default_factory=dict)  # zzp/ez/bv/sv -> J/J+/P/N
    branche: str = ""
    status: str = ""

    @property
    def vervallen(self) -> bool:
        blob = f"{self.status} {self.desc}".lower()
        return "vervallen" in blob

    def applies_to(self, entity: str) -> bool:
        """entity in {zzp, ez, bv, sv}. True if flagged J / J+ / P (basis or uitgebreid)."""
        val = (self.entities.get(entity.lower(), "") or "").strip().upper()
        return val in {"J", "J+", "P"}

    def __str__(self) -> str:
        bits = [self.code]
        if self.nivo:
            bits.append(f"niv{self.nivo}")
        if self.dc:
            bits.append(self.dc)
        if self.nummer:
            bits.append(f"#{self.nummer}")
        head = "  ".join(bits)
        line = f"{head}\n    {self.desc}"
        if self.omslag:
            line += f"\n    omslag → {self.omslag}"
        if self.vervallen:
            line += "\n    ⚠️ VERVALLEN (expired — do not use)"
        return line


# --- Built-in seed (convenience only; NOT the full schema) ----------------------------
# A small map of codes ATTESTED in the skill's source research. niveau 1-3 are grouping
# nodes (not bookable); the niveau-4 entries are the ones confirmed in sources. For any
# code NOT listed here, you MUST consult the official master — do not guess a niveau-4
# tail. Descriptions are indicative; confirm exact spelling against the RGS 3.8 Excel.
_SEED_ROWS = [
    # code, desc, nivo, dc, omslag   (omslag only where a pair is documented)
    # --- niveau 1-2 rubrieken (structure; not bookable) ---
    ("B", "Balans", "1", "", ""),
    ("W", "Winst-en-verliesrekening", "1", "", ""),
    ("BIva", "Immateriële vaste activa", "2", "", ""),
    ("BMva", "Materiële vaste activa", "2", "", ""),
    ("BFva", "Financiële vaste activa", "2", "", ""),
    ("BVrd", "Voorraden", "2", "", ""),
    ("BVor", "Vorderingen", "2", "", ""),
    ("BLim", "Liquide middelen", "2", "", ""),
    ("BEiv", "Eigen vermogen / Kapitaal", "2", "", ""),
    ("BVrz", "Voorzieningen", "2", "", ""),
    ("BLas", "Langlopende schulden", "2", "", ""),
    ("BSch", "Kortlopende schulden", "2", "", ""),
    ("WOmz", "Netto-omzet", "2", "", ""),
    ("WKpr", "Kostprijs van de omzet", "2", "", ""),
    ("WPer", "Lasten uit hoofde van personeelsbeloningen", "2", "", ""),
    ("WAfs", "Afschrijvingen op imm./mat. vaste activa", "2", "", ""),
    ("WBed", "Overige bedrijfskosten", "2", "", ""),
    ("WFbe", "Financiële baten en lasten", "2", "", ""),
    ("WBel", "Belastingen", "2", "", ""),
    ("WNer", "Nettoresultaat", "2", "", ""),
    # --- niveau 3 rubrieken (documented in sources) ---
    ("BVorDeb", "Debiteuren", "3", "", ""),
    ("BVorOva", "Overlopende activa", "3", "", "BSchOpa"),
    ("BSchOpa", "Overlopende passiva", "3", "", "BVorOva"),
    ("BLimKas", "Kasmiddelen", "3", "", ""),
    ("BMvaBeg", "Bedrijfsgebouwen", "3", "", ""),
    # --- niveau 4 bookable accounts (CONFIRMED in sources) ---
    ("BLimKasKas", "Kas", "4", "D", ""),
    ("BMvaBegVvp", "Verkrijgings- of vervaardigingsprijs bedrijfsgebouwen", "4", "D", ""),
    ("WBedAlkOal", "Algemene kosten", "4", "D", ""),          # common general-expenses account
    ("WBedAutOak", "Overige autokosten", "4", "D", ""),
    ("WMfoBelMfo", "Mutatie fiscale oudedagsreserve - belasting (IB; from MoneyBird API docs)", "4", "D", ""),
    # --- niveau 5 mutatie (balance accounts only; example) ---
    ("BMvaBegVvpIna", "Investeringen nieuw aangeschaft bedrijfsgebouwen", "5", "D", ""),
]


def _load_seed() -> dict[str, RgsAccount]:
    out: dict[str, RgsAccount] = {}
    for code, desc, nivo, dc, omslag in _SEED_ROWS:
        out[code] = RgsAccount(code=code, desc=desc, nivo=nivo, dc=dc, omslag=omslag,
                               entities={"bv": "J"})
    return out


# --- Loading ---------------------------------------------------------------------------
def _resolve_headers(headers: list[str], overrides: dict[str, str]) -> dict[str, int]:
    lowered = [h.strip().lower() for h in headers]
    idx: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        if field_name in overrides and overrides[field_name]:
            want = overrides[field_name].strip().lower()
            if want in lowered:
                idx[field_name] = lowered.index(want)
                continue
        for alias in aliases:
            if alias in lowered:
                idx[field_name] = lowered.index(alias)
                break
    return idx


def _row_to_account(row: list[str], idx: dict[str, int]) -> RgsAccount | None:
    def get(name: str) -> str:
        i = idx.get(name)
        if i is None or i >= len(row):
            return ""
        return (row[i] or "").strip()

    code = get("code")
    if not code:
        return None
    return RgsAccount(
        code=code, desc=get("desc"), nummer=get("nummer"), nivo=get("nivo"),
        dc=get("dc"), omslag=get("omslag"), branche=get("branche"), status=get("status"),
        entities={k: get(k) for k in ("zzp", "ez", "bv", "sv")},
    )


def _load_csv(path: str, overrides: dict[str, str]) -> dict[str, RgsAccount]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(fh, delimiter=delim)
        rows = list(reader)
    if not rows:
        return {}
    idx = _resolve_headers(rows[0], overrides)
    if "code" not in idx:
        raise SystemExit(f"Could not find a referentiecode column in {path}. "
                         f"Headers seen: {rows[0]}. Pass --col-code to set it.")
    out: dict[str, RgsAccount] = {}
    for row in rows[1:]:
        acc = _row_to_account(row, idx)
        if acc:
            out[acc.code] = acc
    return out


def _load_xlsx(path: str, overrides: dict[str, str], sheet: str | None) -> dict[str, RgsAccount]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found,import-untyped]
    except ImportError:
        raise SystemExit("Reading .xlsx needs openpyxl (`pip install openpyxl`), or export "
                         "the RGS master to CSV and pass that instead.") from None
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        raise SystemExit(f"Could not open worksheet "
                         f"{repr(sheet) if sheet else '(default/active)'} in {path}. "
                         f"Pass --sheet with a valid tab name.")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return {}
    idx = _resolve_headers(header, overrides)
    if "code" not in idx:
        raise SystemExit(f"No referentiecode column in sheet '{ws.title}'. Headers: {header}. "
                         f"Try --sheet or --col-code.")
    out: dict[str, RgsAccount] = {}
    for raw in rows_iter:
        row = [str(c) if c is not None else "" for c in raw]
        acc = _row_to_account(row, idx)
        if acc:
            out[acc.code] = acc
    return out


def load_rgs(path: str | None, overrides: dict[str, str], sheet: str | None) -> tuple[dict[str, RgsAccount], bool]:
    """Return (accounts_by_code, is_seed)."""
    if not path:
        return _load_seed(), True
    if path.lower().endswith((".xlsx", ".xlsm")):
        return _load_xlsx(path, overrides, sheet), False
    return _load_csv(path, overrides), False


# --- Operations ------------------------------------------------------------------------
def op_validate(db: dict[str, RgsAccount], code: str) -> int:
    acc = db.get(code)
    if not acc:
        print(f"✗ {code}: NOT FOUND in this dataset.")
        return 1
    if acc.vervallen:
        print(f"⚠ {code}: found but VERVALLEN (expired) — do not book to it.\n{acc}")
        return 2
    print(f"✓ {code}: valid.\n{acc}")
    return 0


def op_lookup(db: dict[str, RgsAccount], code: str) -> int:
    acc = db.get(code)
    if not acc:
        # try case-insensitive
        for k, v in db.items():
            if k.lower() == code.lower():
                print(v)
                return 0
        print(f"✗ {code}: not found.")
        return 1
    print(acc)
    return 0


def op_search(db: dict[str, RgsAccount], term: str, entity: str | None,
              nivo: str | None, limit: int) -> int:
    needle = term.lower()
    hits = []
    for acc in db.values():
        if needle not in acc.desc.lower() and needle not in acc.code.lower():
            continue
        if nivo and acc.nivo and acc.nivo != nivo:
            continue
        if entity and acc.entities and not acc.applies_to(entity):
            continue
        if acc.vervallen:
            continue
        hits.append(acc)
    hits.sort(key=lambda a: (a.nivo or "9", a.code))
    if not hits:
        print(f"No matches for '{term}'"
              + (f" (entity={entity})" if entity else "")
              + (f" (nivo={nivo})" if nivo else "") + ".")
        return 1
    for acc in hits[:limit]:
        print(acc)
        print()
    if len(hits) > limit:
        print(f"… {len(hits) - limit} more (raise --limit).")
    return 0


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Look up / validate RGS codes.")
    p.add_argument("--file", help="Official RGS master (.xlsx) or a CSV/TSV export. "
                                   "Omit to use the small built-in seed.")
    p.add_argument("--sheet", help="Worksheet name for .xlsx (e.g. an MKB or Totaal tab).")
    p.add_argument("--validate", metavar="CODE", help="Check a referentiecode exists / is active.")
    p.add_argument("--lookup", metavar="CODE", help="Show one account's details.")
    p.add_argument("--search", metavar="TERM", help="Fuzzy search by description or code.")
    p.add_argument("--entity", choices=["zzp", "ez", "bv", "sv", "ZZP", "EZ", "BV", "SV"],
                   help="Restrict search to codes flagged for this entity type.")
    p.add_argument("--nivo", help="Restrict search to a hierarchy level (4 = bookable).")
    p.add_argument("--limit", type=int, default=25, help="Max search results (default 25).")
    for f in ("code", "desc", "nummer", "nivo", "dc", "omslag"):
        p.add_argument(f"--col-{f}", help=f"Override the '{f}' column header.")
    args = p.parse_args(argv)

    overrides = {f: getattr(args, f"col_{f}") for f in ("code", "desc", "nummer", "nivo", "dc", "omslag")}
    db, is_seed = load_rgs(args.file, overrides, args.sheet)
    if is_seed:
        print("(i) Using built-in SEED dataset (~20 common BV codes). Not authoritative - "
              "download the official RGS 3.8 Excel and pass --file for real work.\n",
              file=sys.stderr)
    else:
        print(f"Loaded {len(db)} RGS codes from {args.file}.\n", file=sys.stderr)

    if args.validate:
        return op_validate(db, args.validate)
    if args.lookup:
        return op_lookup(db, args.lookup)
    if args.search:
        return op_search(db, args.search, args.entity, args.nivo, args.limit)
    p.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
